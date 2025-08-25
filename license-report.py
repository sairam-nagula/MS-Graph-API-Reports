import pytz
import pandas as pd
from dotenv import load_dotenv
from email.message import EmailMessage
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from msal import ConfidentialClientApplication
import os, io, csv, pathlib, requests, smtplib

# Configurations
GRAPH = "https://graph.microsoft.com"
API = GRAPH + "/v1.0"
PERIOD_DAYS = 30
STALE_DAYS = 45
OUTFILE = pathlib.Path("m365_license_health.xlsx")
est = pytz.timezone("US/Eastern")

UNIT_COST_DICTIONARY = {
    "ATP_ENTERPRISE": 3.00,
    "Clipchamp_Standard": 7.00,
    "DYN365_BUSCENTRAL_ESSENTIAL": 70.00,
    "ENTERPRISEPACK": 36.00,
    "EXCHANGEDESKLESS": 4.00,
    "EXCHANGEENTERPRISE": 8.00,
    "EXCHANGESTANDARD": 4.00,
    "MCOMEETADV": 2.00,
    "MCOTEAMS_ESSENTIALS": 4.00,
    "Microsoft_365_Copilot": 30.00,
    "Microsoft_365_E3_(no_Teams)": 33.00,
    "Microsoft_Teams_Enterprise_New": 7.00,
    "Microsoft_Teams_Exploratory_Dept": 0.00,
    "Microsoft_Teams_Premium": 7.00,
    "Microsoft_Teams_Rooms_Basic": 0.00,
    "Microsoft_Teams_Rooms_Pro": 40.00,
    "PBI_PREMIUM_PER_USER": 20.00,
    "POWERAPPS_DEV": 10.00,
    "POWER_BI_PRO": 10.00,
    "PROJECTPROFESSIONAL": 30.00,
    "SPE_E3": 36.00,
    "SPE_E5": 57.00,
    "THREAT_INTELLIGENCE": 2.00,
    "Teams_Phone_with_domestic_and_international_calling": 15.00,
    "Teams_Premium_(for_Departments)": 7.00,
}


def get_graph_api_access_token():
    """
    Authenticate to Azure and obtain an access token for the Graph API.

    The `ConfidentialClientApplication` class is used to authenticate to Azure and
    obtain an access token for the Graph API.

    The access token is returned as a string.
    """
    
    load_dotenv()
    tenant = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    app = ConfidentialClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant}",
        client_credential=client_secret
    )
    token = app.acquire_token_for_client(scopes=[f"{GRAPH}/.default"])
    return token["access_token"]


def graph_api_get_request(url, token, stream=False):
    """
    Send a GET request to the Graph API.

    The Graph API URL is taken from the `url` argument.

    The response is returned as a `requests.Response` object.
    """
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "text/csv" if "/reports/" in url else "application/json"
    }
    return requests.get(url, headers=headers, stream=stream)


def pagination_helper(url, token):
    """
    Iterate over all pages of a Graph API query.

    The Graph API URL is taken from the `url` argument.

    The `token` argument is the access token for the Graph API.

    The function yields each item in the response, one at a time.

    The function will continue to make requests until all pages are exhausted.

    The function will return immediately if the `url` argument is `None`.
    """
    while url:
        data = graph_api_get_request(url, token).json()
        for item in data.get("value", []):
            yield item
        url = data.get("@odata.nextLink")


def parse_date_any(date):
 
    """
    Parse a date string, or return NaT (Not a Time) if not parseable.

    The string is first stripped of any whitespace and then any zero-width
    spaces, non-breaking spaces, or "byte order mark" characters are replaced
    with regular spaces. The string is then parsed using pandas to_datetime
    with errors="coerce", which means that if the string cannot be parsed, the
    function will return NaT. The timezone is set to UTC.

    Parameters
    ----------
    date : str or None
        The string to be parsed.

    Returns
    -------
    pd.Timestamp or pd.NaT
        The parsed date or NaT if the string was not parseable.

    """

    if date is None:
        return pd.NaT
    
    date_string = str(date).strip().replace("\u00A0"," ").replace("\u200b","").replace("\ufeff","")
    
    if not date_string or date_string.lower() == "nan":
        return pd.NaT
    
    date_string = date_string.replace("Z","").replace(".000","")
    
    return pd.to_datetime(date_string, errors="coerce", utc=True)


def get_licenses(token):
    """
    Get all subscribed skus with their enabled, consumed, and remaining units

    The function takes an access token for the Graph API as a parameter.

    The function returns a pandas DataFrame with the following columns:

    - skuId
    - skuPartNumber
    - total_enabled
    - consumed
    - remaining
    - warning
    - suspended
    - estMonthlyCost

    The function will return an empty DataFrame if the request fails.

    Parameters
    ----------
    token : str
        The access token for the Graph API.

    Returns
    -------
    pd.DataFrame
        A DataFrame with the subscribed skus and their enabled, consumed, and remaining units.
    """
    
    url = f"{API}/subscribedSkus?$select=skuId,skuPartNumber,prepaidUnits,consumedUnits"
    
    response = graph_api_get_request(url, token).json()
    
    rows = []
    
    for sku in response.get("value", []):
        prepaid = sku.get("prepaidUnits", {})
        enabled = prepaid.get("enabled", 0)
        consumed = sku.get("consumedUnits", 0)
        remaining = max(enabled - consumed, 0)
        part = sku["skuPartNumber"]
        est_cost = round(consumed * UNIT_COST_DICTIONARY.get(part, 0), 2)
        
        rows.append({
            "skuId": sku["skuId"],
            "skuPartNumber": part,
            "total_enabled": enabled,
            "consumed": consumed,
            "remaining": remaining,
            "warning": prepaid.get("warning", 0),
            "suspended": prepaid.get("suspended", 0),
            "estMonthlyCost": est_cost
        })
        
    return pd.DataFrame(rows)


def get_all_users(token):  
    """
    Get all users with their details

    The function takes an access token for the Graph API as a parameter.

    The function returns a pandas DataFrame with the following columns:

    - id
    - displayName
    - userPrincipalName (UPN)
    - mail
    - accountEnabled
    - userType
    - createdDateTime
    - assignedLicenses (a list of skuIds)
    - assignedPlans (a list of plans)

    The function will return an empty DataFrame if the request fails.

    Parameters
    ----------
    token : str
        The access token for the Graph API.

    Returns
    -------
    pd.DataFrame
        A DataFrame with the users and their details.
    """

    fields = ",".join([
        "id", "displayName", "userPrincipalName", "mail", 
        "accountEnabled", "userType", "createdDateTime", 
        "assignedLicenses", "assignedPlans"
    ])
    
    url = f"{API}/users?$select={fields}&$top=999"
    
    users = []
    
    for user in pagination_helper(url, token):
        users.append({
            "id": user["id"],
            "displayName": user.get("displayName", ""),
            "UPN": user.get("userPrincipalName", ""),
            "mail": user.get("mail", ""),
            "accountEnabled": user.get("accountEnabled", True),
            "userType": user.get("userType", "Member"),
            "createdDateTime": user.get("createdDateTime", ""),
            "assignedLicenses": [lic.get("skuId") for lic in user.get("assignedLicenses", [])],
            "assignedPlans": user.get("assignedPlans", [])
        })
        
    return pd.DataFrame(users)


def get_users_activity_status(token):
    """
    Get the activity status of users in the tenant.

    The function takes an access token for the Graph API as a parameter.

    The function returns a pandas DataFrame with two columns: UPN_lower and LastActivityDate.
    The UPN_lower column contains the user principal name in lower case.
    The LastActivityDate column contains the date of the last activity of the user.

    The function will return an empty DataFrame if the request fails.

    Parameters
    ----------
    token : str
        The access token for the Graph API.

    Returns
    -------
    pd.DataFrame
        A DataFrame with the activity status of users.
    """
    
    url = f"{API}/reports/getOffice365ActiveUserDetail(period='D{PERIOD_DAYS}')"
    
    response = graph_api_get_request(url, token, stream=True).content.decode("utf-8-sig", errors="ignore")
    
    reader = csv.DictReader(io.StringIO(response))
    
    user_activity = []
    
    for row in reader:
        upn = (
            row.get("User Principal Name") or 
            row.get("UPN") or 
            row.get("User Id") or ""
        ).strip().lower()
        
        if not upn or upn.startswith(("user ", "hidden", "redacted")):
            continue
        activity = row.get("Last Activity Date") or row.get("Last Activity Date (UTC)") or row.get("Report Refresh Date")
        user_activity.append({"UPN_lower": upn, "LastActivityDate": activity})
    
    return pd.DataFrame(user_activity)


def df_to_email_html(df):
    """
    Convert a pandas DataFrame to an HTML string suitable for embedding in an email body.

    The function takes a pandas DataFrame as a parameter.

    The function returns an HTML string that can be embedded in an email body.

    The function will return an empty string if the DataFrame is empty.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame to convert to HTML.

    Returns
    -------
    str
        The HTML string representation of the DataFrame.
    """
    
    style = (
        "border-collapse:collapse;font-family:Segoe UI,Arial,sans-serif;font-size:12px;"
    )
    
    th = "background:#f2f2f2;border:1px solid #ddd;padding:6px 8px;text-align:left;"
    
    td = "border:1px solid #ddd;padding:6px 8px;text-align:left;"

    # build rows manually for clean control
    headers = "".join(f"<th style='{th}'>{h}</th>" for h in df.columns)
    
    body_rows = []
    
    for _, row in df.iterrows():
        tds = "".join(f"<td style='{td}'>{'' if pd.isna(v) else v}</td>" for v in row)
        body_rows.append(f"<tr>{tds}</tr>")
    
    body = "\n".join(body_rows)
    
    return f"<table style='{style}'><thead><tr>{headers}</tr></thead><tbody>{body}</tbody></table>"


def send_email(attachment_path, overview_html):
    
    """
    Send an email with the report to the recipients.

    This function sends an email with the report as an attachment, and an HTML body
    that includes the Overview table.

    Parameters
    ----------
    attachment_path : str
        The path to the Excel file containing the report.
    overview_html : str
        The HTML content of the Overview table.

    Returns
    -------
    None
    """
    
    now = datetime.now(pytz.utc).astimezone(est).strftime("%m-%d-%Y %I:%M%p")
    
    msg = EmailMessage()
    
    msg["Subject"] = f"Microsoft License Report - {now} EST"
    msg["From"] = os.environ["OFFICE_365_USERNAME"]
    msg["To"] = os.environ.get("EMAIL_RECIPIENTS", "").split(",")

    # HTML body with the Overview table embedded
    msg.add_alternative(f"""
        <html><body>
        <p>Hello,</p>
        <p>Here is the overview for this months Microsoft Licenses report:</p>
        {overview_html}
        <p>The detailed tabs are in the attached file.</p>
        </body></html>
    """, subtype="html")

    with open(attachment_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=pathlib.Path(attachment_path).name
        )

    smtp = smtplib.SMTP("smtp.office365.com", 587)
    smtp.starttls()
    smtp.login(os.environ["OFFICE_365_USERNAME"], os.environ["OFFICE_365_PASSWORD"])
    smtp.send_message(msg)
    smtp.quit()
    
    print("Email sent.")


def process_and_export_data(skus_df, users_df, activity_df, outfile_path):
    """
    Process and export data from the Microsoft Graph API to an Excel file.

    The function takes in the dataframes from the Microsoft Graph API, processes them
    and exports the data to an Excel file with the specified filename.

    Parameters
    ----------
    skus_df : pandas.DataFrame
        The dataframe of all available SKUs
    users_df : pandas.DataFrame
        The dataframe of all users
    act_df : pandas.DataFrame
        The dataframe of all user activity
    outfile_path : str
        The path to the output Excel file

    Returns
    -------
    overview_html : str
        The HTML for the Overview sheet as a string
        """

    # Data cleaning and preparation
    activity_df["LastActivityDate"] = pd.to_datetime(activity_df["LastActivityDate"], errors="coerce", utc=True)
    users_df["UPN_lower"] = users_df["UPN"].fillna("").astype(str).str.strip().str.lower()
    activity_df["UPN_lower"] = activity_df["UPN_lower"].fillna("").astype(str).str.strip().str.lower()
    merged_df = users_df.merge(activity_df, how="left", on="UPN_lower")

    # Timezone adjustments
    merged_df["LastActivityDate"] = merged_df["LastActivityDate"].dt.tz_convert("US/Eastern").dt.tz_localize(None).dt.date
    merged_df["createdDateTime"] = pd.to_datetime(merged_df["createdDateTime"], errors="coerce", utc=True)
    merged_df["createdDateTime"] = merged_df["createdDateTime"].dt.tz_convert("US/Eastern").dt.tz_localize(None).dt.date

    # Inactive flag
    inactivity_cutoff = (datetime.now(est) - timedelta(days=PERIOD_DAYS)).date()
    merged_df["Inactive30d"] = merged_df["LastActivityDate"].apply(lambda ts: True if pd.isna(ts) else ts <= inactivity_cutoff)


    # Build license string from assignedLicenses
    license_map = dict(zip(skus_df["skuId"], skus_df["skuPartNumber"]))
    merged_df["licenses"] = merged_df["assignedLicenses"].apply(
        lambda arr: ";".join(sorted({license_map.get(x, str(x)) for x in (arr or [])}))
    )
    merged_df["hasLicense"] = merged_df["licenses"].str.len() > 0

    # Paid license checks
    paid_skus = {k for k, v in UNIT_COST_DICTIONARY.items() if (v or 0) > 0}
    exclude_for_disabled = {"EXCHANGEENTERPRISE"}
    paid_excl = paid_skus.difference(exclude_for_disabled)

    def has_paid(lics): return any(sku in paid_skus for sku in lics.split(";") if sku)
    def has_paid_excl(lics): return any(sku in paid_excl for sku in lics.split(";") if sku)

    # Add paid license flags
    merged_df["hasPaidLicense"] = merged_df["licenses"].apply(has_paid)
    merged_df["hasPaidLicenseExclExchEnt"] = merged_df["licenses"].apply(has_paid_excl)

    # Actionables
    actionable = pd.DataFrame()

    to_do1 = merged_df[(merged_df["hasLicense"]) & (merged_df["Inactive30d"]) & (merged_df["hasPaidLicense"])].copy()
    to_do1["Reason"] = "Licensed but no activity in last 30d"
    actionable = pd.concat([actionable, to_do1], ignore_index=True)

    to_do2 = merged_df[
        (merged_df["hasLicense"]) &
        (merged_df["hasPaidLicenseExclExchEnt"]) &
        (merged_df["accountEnabled"] == False)
    ].copy()
    to_do2["Reason"] = "Disabled account has licenses"
    actionable = pd.concat([actionable, to_do2], ignore_index=True)

    # SKU usage stats
    sku_stats = []
    for _, row in merged_df.iterrows():
        if not row["hasLicense"]:
            continue
        active = not row["Inactive30d"]
        for sku in filter(None, row["licenses"].split(";")):
            sku_stats.append((sku, 1, 1 if active else 0))
    if sku_stats:
        sku_frame = pd.DataFrame(sku_stats, columns=["skuPartNumber", "licensedUsers", "activeUsers30d"])
        util = sku_frame.groupby("skuPartNumber", as_index=False).sum()
    else:
        util = pd.DataFrame(columns=["skuPartNumber", "licensedUsers", "activeUsers30d"])
    util["utilizationPct30d"] = (util["activeUsers30d"] / util["licensedUsers"]).round(4) * 100.0

    # Merge utilization into SKU summary
    sku_summary = skus_df.merge(util, how="left", on="skuPartNumber")
    
    sku_summary[["licensedUsers", "activeUsers30d", "utilizationPct30d"]] = sku_summary[
        ["licensedUsers", "activeUsers30d", "utilizationPct30d"]
    ].fillna(0)

    # Paid SKUs
    sku_summary["unitCost"] = sku_summary["skuPartNumber"].map(UNIT_COST_DICTIONARY).fillna(0.0)
    has_enabled = sku_summary["total_enabled"] > 0
    ends_bulk = sku_summary["total_enabled"].astype(str).str.endswith(("00", "000"))
    is_paid = sku_summary["unitCost"] > 0
    paid_summary = sku_summary[has_enabled & (~ends_bulk) & is_paid].copy()
    paid_summary["estMonthlyCost_num"] = pd.to_numeric(paid_summary["estMonthlyCost"], errors="coerce").fillna(0.0)
    cost_summary = paid_summary[paid_summary["estMonthlyCost_num"] > 0].copy()

    # KPIs
    total_licensed = int(util["licensedUsers"].sum()) if not util.empty else 0
    total_active = int(util["activeUsers30d"].sum()) if not util.empty else 0
    util_pct = round((total_active / total_licensed) * 100, 2) if total_licensed else 0.0

    kpis = pd.DataFrame([
        {"Metric": "Report Timestamp", "Value": datetime.utcnow().strftime("%m-%d-%y")},
        {"Metric": "Licensed Users ", "Value": total_licensed},
        {"Metric": f"Users Active in Last {PERIOD_DAYS} days", "Value": total_active},
        {"Metric": f"Overall % Utilization {PERIOD_DAYS} days", "Value": f"{util_pct}%"},
        {"Metric": "To-Do: Licensed but inactive 30 days", "Value": int((to_do1["UPN"].nunique() if not to_do1.empty else 0))},
        {"Metric": "To-Do: Disabled & licensed", "Value": int((to_do2["UPN"].nunique() if not to_do2.empty else 0))},
    ])

    # Output columns
    users_out_cols = [
        "displayName", "UPN", "userType", "accountEnabled", "createdDateTime",
        "licenses", "LastActivityDate", "Inactive30d"
    ]
    
    actionable_cols = [
        "Reason", "displayName", "UPN", "userType", "accountEnabled", "createdDateTime",
        "licenses", "LastActivityDate"
    ]
    
    sku_cols = [
        "skuPartNumber", "total_enabled", "remaining", "licensedUsers",
        "activeUsers30d", "utilizationPct30d", "suspended", "estMonthlyCost"
    ]
    
    CUSTOM_SKU_HEADERS = {
        "skuPartNumber": "License SKU's",
        "total_enabled": "Purchased",
        "remaining": "Remaining",
        "licensedUsers": "Assigned",
        "activeUsers30d": "Active (30d)",
        "utilizationPct30d": "Utilization (30d, %)",
        "suspended": "Suspended",
        "estMonthlyCost": "Est. Monthly Cost ($)"
    }

    # Write Excel file
    with pd.ExcelWriter(outfile_path, engine="openpyxl") as xw:
        kpis.to_excel(xw, sheet_name="Overview", index=False)
        
        cost_summary[sku_cols].sort_values("skuPartNumber").rename(
            columns=CUSTOM_SKU_HEADERS
        ).to_excel(xw, sheet_name="SKU_Summary", index=False)
        
        actionable[actionable_cols].drop_duplicates(
            subset=["Reason", "UPN"]
        ).sort_values(["Reason", "UPN"]).to_excel(xw, sheet_name="Actionable", index=False)
        
        merged_df[users_out_cols].sort_values("UPN").to_excel(xw, sheet_name="Users", index=False)
        

    # Write Excel file 
    with pd.ExcelWriter(outfile_path, engine="openpyxl") as xw:

        cost_summary[sku_cols].sort_values("skuPartNumber").rename(
            columns=CUSTOM_SKU_HEADERS
        ).to_excel(xw, sheet_name="SKU_Summary", index=False)

        actionable[actionable_cols].drop_duplicates(
            subset=["Reason", "UPN"]
        ).sort_values(["Reason", "UPN"]).to_excel(xw, sheet_name="Actionable", index=False)

        merged_df[users_out_cols].sort_values("UPN").to_excel(xw, sheet_name="Users", index=False)
         
        # Column width adjustments
        wb = xw.book
        
        for sheet in ["SKU_Summary"]:
            ws = wb[sheet]
            ws.column_dimensions[get_column_letter(1)].width = 40   
            
        for sheet in ["Users"]:
            ws = wb[sheet]
            ws.column_dimensions[get_column_letter(1)].width = 30   
            ws.column_dimensions[get_column_letter(2)].width = 25   
            ws.column_dimensions[get_column_letter(5)].width = 15   
            ws.column_dimensions[get_column_letter(7)].width = 15   
            ws.column_dimensions[get_column_letter(6)].width = 30
            
        for sheet in ["Actionable"]:
            ws = wb[sheet]
            ws.column_dimensions[get_column_letter(1)].width = 30
            ws.column_dimensions[get_column_letter(2)].width = 25   
            ws.column_dimensions[get_column_letter(6)].width = 15     
            ws.column_dimensions[get_column_letter(7)].width = 30   
            ws.column_dimensions[get_column_letter(8)].width = 15   
            
        overview_html = df_to_email_html(kpis)
        
    return overview_html


def main():
    OUTFILE.parent.mkdir(parents=True, exist_ok=True)
    token = get_graph_api_access_token()

    skus = get_licenses(token)
    users = get_all_users(token)
    activity = get_users_activity_status(token)

    overview_html = process_and_export_data(skus, users, activity, OUTFILE)

    send_email(str(OUTFILE), overview_html)


if __name__ == "__main__":
    main()
